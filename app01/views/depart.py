from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from app01.utils.pagination import Pagination
from openpyxl import load_workbook


def depart_list(request):
    # 部门列表
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        'departments': page_object.page_queryset,
        'page_string': page_object.html()
    }
    return render(request, 'depart_list.html', context)


def depart_add(request):
    # 添加部门
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户POST提交过来的数据
    department_name = request.POST.get("department_name")
    # 保存至数据库
    models.Department.objects.create(title=department_name)
    # 重定向回到列表页面
    return redirect('/depart/list/')


def depart_delete(request, nid):
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')


def depart_edit(request, nid):
    # 修改部门
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {"row_object": row_object})
    # 获取用户提交的标题
    title = request.POST.get("title")
    # 根据ID找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向回部门列表
    return redirect('/depart/list/')


def depart_multi(request):
    file_obj = request.FILES.get('exc')
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')
